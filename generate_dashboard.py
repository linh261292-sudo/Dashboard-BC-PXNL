#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_dashboard.py
======================
Tu dong cap nhat "index.html" (bao cao lanh dao - kho than NMND Song Hau 1)
tu file Excel nguon "So do luu kho than.xlsx" nam CUNG THU MUC voi script nay.

Cach dung:
    python generate_dashboard.py

Yeu cau:
    pip install openpyxl

Script nay trich xuat du lieu tho tu cac sheet Excel (2 sheet BAT BUOC: Sheet1,
Bao cao san luong; va cac sheet TUY CHON bo sung: Tau OG Indonesia-VTau, Tau SB
tai VTau- Cang SH1, Theo doi KL HD giao nhan, Ke hoach tau PO trong thang) va
ghi de vao khoi "var RAW_DATA = {...};" trong index.html. Toan bo tinh toan/
dien giai (KPI, bang, bieu do, tau me OG duoc suy ra tu sheet Indonesia-VTau...)
nam trong JS (computeReport/renderText) o index.html va se tu dong chay lai
moi khi trang duoc mo - script Python nay khong can biet gi ve cach hien thi,
chi can trich xuat dung so lieu.

An toan: neu thieu 1 trong 2 sheet BAT BUOC, hoac thieu file, hoac cau truc
index.html bi doi, script se dung lai va KHONG ghi de index.html, de tranh
lam hong dashboard dang chay. Cac sheet TUY CHON neu thieu/doi ten chi bo qua
phan du lieu tuong ung (mang rong []) va ghi CANH BAO vao log, khong dung script.
"""
import json
import os
import re
import sys

# QUAN TRONG (loi thuc te da gap trong log tu dong chay tren may Windows): khi script chay
# qua Task Scheduler va bi redirect output vao file log (">> log.txt"), Python/Windows mac
# dinh dung code page cua may (thuong la cp1252), khong phai UTF-8 — bat ky ky tu tieng Viet
# co dau nao (a, e, o, u, d...) trong thong bao loi se lam print() nem UnicodeEncodeError,
# khien chinh buoc BAO LOI lai bi crash va che mat thong bao loi that su trong log. Ep stdout/
# stderr sang UTF-8 ngay tu dau de tranh hoan toan loi nay.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(HERE, "So do luu kho than.xlsx")
HTML_PATH = os.path.join(HERE, "index.html")

# 2 sheet BAT BUOC — thieu 1 trong 2 la dung script, khong ghi de index.html.
SHEET_STOCK = "Sheet1"
SHEET_PRODUCTION = "Báo cáo sản lượng"

# Cac sheet duoi day la TUY CHON (khong bat buoc) — neu chua co / bi doi ten, script
# van chay binh thuong va chi bo qua phan du lieu tuong ung (mang rong []), KHONG
# lam dung toan bo script, vi day la du lieu bo sung cho cac slide tau than/PO.
SHEET_OG_INDO = "Tau OG Indonesia-VTau"
SHEET_SB = "Tàu SB tại VTau- Cảng SH1"
SHEET_HD_TRACKING = "Theo dõi KL HĐ giao nhận"
SHEET_PO = "Kế hoạch tàu PO trong tháng"
# 4 sheet CHI TIET tung "Lô" boc do theo TUNG hop dong rieng (khac voi SHEET_HD_TRACKING o tren -
# sheet do chi co 1 dong tong hop/hop dong; 4 sheet nay liet ke tung tau TS/SB rieng le trong tung
# lo hang) - cung la TUY CHON, dung cho bang "thong tin tau SB" o Slide 05.
SHEET_HD_DETAIL = {
    "HĐ 17": "HĐ 17",
    "HĐ 24": "HĐ 24",
    "HĐ 25": "HĐ 25",
    "HĐ 26": "HĐ 26",
}


def log(msg):
    # phong ngua kep: du reconfigure() o tren co that bai vi ly do gi, log() van khong duoc
    # phep tu crash — neu in UTF-8 loi thi thu lai voi errors='replace' thay vi nem exception.
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        try:
            enc = sys.stdout.encoding or "ascii"
            print(msg.encode(enc, errors="replace").decode(enc), flush=True)
        except Exception:
            print(msg.encode("ascii", errors="replace").decode("ascii"), flush=True)


def die(msg):
    log("LOI: " + msg)
    sys.exit(1)


def fmt_date(v):
    """Dinh dang 1 gia tri ngay Excel (datetime) thanh chuoi 'dd-mm-yy'; neu khong
    phai datetime (vd da la text san) thi tra ve nguyen chuoi."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d-%m-%y")
    return str(v)


def extract_stock_rows(ws):
    """Cac lo than hien dang o bai: dang luu kho binh thuong, dang nhap, hoac dang cap.
    Trang thai (cot 'status') xac dinh theo cot NgayCapHet (cot L/12, UU TIEN cao nhat)
    va cot TrangThaiKho (cot O/15):
      - NgayCapHet CO gia tri (khac rong) -> da cap het / da xu ly xong -> KHONG con o bai
        nua -> bo qua hoan toan, bat ke TrangThaiKho la gi (tranh hien thi nham cac lo
        da xong tu lau nhung con sot co TrangThaiKho='Nhap'/'Cap' cu chua don dep).
      - NgayCapHet rong VA TrangThaiKho == 'Nhap' -> status 'nhap' (dang nhap -> do)
      - NgayCapHet rong VA TrangThaiKho == 'Cap'  -> status 'cap'  (dang cap  -> vang)
      - NgayCapHet rong VA TrangThaiKho rong       -> status 'luukho' (dang luu kho -> xanh la)
    """
    rows = []
    r = 2
    empty_streak = 0
    while empty_streak < 30 and r < 5000:
        ten_tau = ws.cell(r, 1).value
        if ten_tau is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0

        trang_thai_raw = ws.cell(r, 15).value
        trang_thai = str(trang_thai_raw).strip() if trang_thai_raw is not None else ""
        ngay_cap_het = ws.cell(r, 12).value

        if ngay_cap_het is not None:
            status = None
        elif trang_thai == "Nhap":
            status = "nhap"
        elif trang_thai == "Cap":
            status = "cap"
        else:
            status = "luukho"

        if status:
            yard = ws.cell(r, 4).value
            zone = ws.cell(r, 5).value
            layer = ws.cell(r, 6).value
            segment = ws.cell(r, 7).value
            ngay_nhap = ws.cell(r, 8).value
            kl_nhap = ws.cell(r, 9).value
            tau_sb = ws.cell(r, 2).value  # cot TauSB: ten tau/sa lan con da trung chuyen dung lo nay vao bai
            if yard and zone and layer is not None and segment and ngay_nhap and kl_nhap is not None:
                rows.append({
                    "yard": str(yard).strip(),
                    "zone": str(zone).strip(),
                    "layer": int(layer),
                    "segment": str(segment).strip(),
                    "vessel": str(ten_tau).strip(),
                    "sbVessel": str(tau_sb).strip() if tau_sb else None,
                    "date": ngay_nhap.strftime("%d/%m/%Y"),
                    "qty": round(float(kl_nhap), 2),
                    "status": status,
                })
        r += 1
    return rows


def extract_pile_stock_rows(ws):
    """Nguon du lieu RIENG cho slide 3D 'Bo tri luu tru kho than' (phien ban 2, chen ngay
    sau slide 'Bo tri dong than luu kho' cu). Dung LAI DUNG logic uu tien NgayCapHet nhu
    extract_stock_rows() o tren (NgayCapHet CO gia tri bat ke dinh dang/gia tri gi -> lo
    da cap het, KHONG con o bai -> loai bo hoan toan, o do se hien "Dang trong" tren so do
    3D vi khong con block nao duoc ve tai do). Khac voi extract_stock_rows() o 2 diem:
      1. Chuan hoa ten yard (trim khoang trang thua, vd 'Đống 2.1 ' -> 'Đống 2.1') vi du
         lieu tho co nhieu dong bi du/thieu dau cach sau ten dong.
      2. Loai han yard 'Bypass' (than trung chuyen thang, khong thuoc 4 dong vat ly hien
         tren so do bai chua) - slide nay chi ve dung 4 dong: Đống 1, Đống 2.1, Đống 2.2,
         Đống 3.
    """
    rows = []
    r = 2
    empty_streak = 0
    valid_yards = {"Đống 1", "Đống 2.1", "Đống 2.2", "Đống 3"}
    while empty_streak < 30 and r < 5000:
        ten_tau = ws.cell(r, 1).value
        if ten_tau is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0

        ngay_cap_het = ws.cell(r, 12).value
        yard_raw = ws.cell(r, 4).value
        yard = str(yard_raw).strip() if yard_raw else None
        zone = ws.cell(r, 5).value
        layer = ws.cell(r, 6).value
        segment = ws.cell(r, 7).value
        ngay_nhap = ws.cell(r, 8).value
        kl_nhap = ws.cell(r, 9).value
        tau_sb = ws.cell(r, 2).value
        trang_thai_raw = ws.cell(r, 15).value
        trang_thai = str(trang_thai_raw).strip() if trang_thai_raw is not None else ""

        if (ngay_cap_het is None and yard in valid_yards and zone and layer is not None
                and segment and ngay_nhap and kl_nhap is not None):
            if trang_thai == "Nhap":
                status = "nhap"
            elif trang_thai == "Cap":
                status = "cap"
            else:
                status = "luukho"
            rows.append({
                "yard": yard,
                "zone": str(zone).strip(),
                "layer": int(layer),
                "segment": str(segment).strip(),
                "vessel": str(ten_tau).strip(),
                "sbVessel": str(tau_sb).strip() if tau_sb else None,
                "date": ngay_nhap.strftime("%d/%m/%Y") if hasattr(ngay_nhap, "strftime") else str(ngay_nhap),
                "qty": round(float(kl_nhap), 2),
                "status": status,
            })
        r += 1
    return rows


def extract_month_days(ws):
    """Doc toan bo cac ngay co du lieu ton kho (cot F/6), roi chi lay cac
    ngay thuoc THANG cua ngay cuoi cung co du lieu (thang bao cao hien tai)."""
    all_rows = []
    r = 3
    empty_streak = 0
    while empty_streak < 60 and r < 3000:
        d = ws.cell(r, 1).value
        if d is None:
            empty_streak += 1
            r += 1
            continue
        tonkho = ws.cell(r, 6).value
        if tonkho is not None:
            nhap = ws.cell(r, 4).value or 0
            tieuthu = ws.cell(r, 5).value or 0
            all_rows.append((d, float(nhap), float(tieuthu), float(tonkho)))
            empty_streak = 0
        else:
            empty_streak += 1
        r += 1
    if not all_rows:
        return None, None, None, []
    last_date = all_rows[-1][0]
    month, year = last_date.month, last_date.year
    month_days = [
        {"day": d.day, "nhap": round(n, 2), "tieuthu": round(t, 2), "tonkho": round(tk, 2)}
        for (d, n, t, tk) in all_rows if d.month == month and d.year == year
    ]
    month_days.sort(key=lambda x: x["day"])
    return last_date.strftime("%d/%m/%Y"), month, year, month_days


def extract_og_indo_vtau(ws):
    """Sheet 'Tau OG Indonesia-VTau': DUY NHAT nguon du lieu tau me OG — vua theo
    doi vi tri/hanh trinh (Indonesia -> Go Gia), vua la nguon tien do (KL Nhap vao
    Kho SH1) dung de suy ra bang "Tau me OG dang giao nhan" ben JS (computeReport).
    Cot: B=ten, C=vi tri/ETA (Indonesia / Dang den Indonesia / Indonesia-Go Gia /
    Go Gia), D=ngay roi cang Indonesia, E=ngay du kien den Go Gia, F=hop dong,
    G=OG so, H=khoi luong hang, I=da xep hang tai Indonesia, J=KL xep con lai,
    K=KL da nhap vao kho SH1 (tuy chon, co the trong voi tau chua ve den kho)."""
    rows = []
    r = 2
    empty_streak = 0
    while empty_streak < 15 and r < 500:
        name = ws.cell(r, 2).value
        if name is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        position = ws.cell(r, 3).value
        depart = ws.cell(r, 4).value
        eta = ws.cell(r, 5).value
        contract = ws.cell(r, 6).value
        code = ws.cell(r, 7).value
        kl_hang = ws.cell(r, 8).value
        da_xep = ws.cell(r, 9).value
        con_lai = ws.cell(r, 10).value
        kl_nhap_kho = ws.cell(r, 11).value
        if kl_hang is not None:
            da_xep_v = float(da_xep) if da_xep is not None else 0.0
            con_lai_v = float(con_lai) if con_lai is not None else max(float(kl_hang) - da_xep_v, 0.0)
            kl_nhap_kho_v = float(kl_nhap_kho) if kl_nhap_kho is not None else 0.0
            rows.append({
                "name": str(name).strip(),
                "position": str(position).strip() if position else "",
                "departIndo": fmt_date(depart),
                "etaVtau": fmt_date(eta),
                "contract": str(contract).strip() if contract else "",
                "code": str(code).strip() if code else "",
                "klHang": round(float(kl_hang), 2),
                "daXepIndo": round(da_xep_v, 2),
                "conLai": round(con_lai_v, 2),
                "klNhapKho": round(kl_nhap_kho_v, 2),
            })
        r += 1
    return rows


def extract_sb_vessels(ws):
    """Sheet 'Tàu SB tại VTau- Cảng SH1': chi con tau/sa lan (SB) trung chuyen than
    tu tau me OG vao bo (tau me OG gio lay tu sheet Tau OG Indonesia-VTau, KHONG
    con trong sheet nay nua). Cot: B=ten, C=hop dong, D=OG so (ma tau me de doi
    chieu), E=loai (luon la 'SB'), F=vi tri/ETA (Cai Cui / Song Hau 1 / Go Gia /
    Go Gia-Cai Cui), G=khoi luong hang, H=da boc do, I=con lai."""
    rows = []
    r = 2
    empty_streak = 0
    while empty_streak < 15 and r < 500:
        name = ws.cell(r, 2).value
        if name is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        contract = ws.cell(r, 3).value
        og_so = ws.cell(r, 4).value
        kl = ws.cell(r, 7).value
        done = ws.cell(r, 8).value
        remain = ws.cell(r, 9).value
        vitri = ws.cell(r, 6).value
        if kl is not None:
            done_v = float(done) if done is not None else 0.0
            remain_v = float(remain) if remain is not None else max(float(kl) - done_v, 0.0)
            rows.append({
                "name": str(name).strip(),
                "contract": str(contract).strip() if contract else "",
                "parentCode": str(og_so).strip() if og_so else None,
                "position": str(vitri).strip() if vitri else "",
                "klNor": round(float(kl), 2),
                "done": round(done_v, 2),
                "remain": round(remain_v, 2),
            })
        r += 1
    return rows


def extract_hd_tracking(ws):
    """Sheet 'Theo dõi KL HĐ giao nhận' (cau truc hien tai: KHONG con cot rieng "Thoi
    gian ket thuc giao" — thay vao do la 2 cot du bao o cuoi bang). Cot: B=ten HD,
    C=thoi gian bat dau giao, D=KL Giao, E=KL da xep, F=KL da do vao kho, G=con lai,
    H=thoi gian du kien Nhap het (dung lam gia tri hien thi o cot "Ket thuc" tren bang
    bao cao), I=thoi gian du kien Dot het (chua dung toi, JS tu tinh rieng tu toc do
    tieu thu binh quan)."""
    rows = []
    r = 2
    empty_streak = 0
    while empty_streak < 15 and r < 500:
        name = ws.cell(r, 2).value
        if name is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        start = ws.cell(r, 3).value
        kl_giao = ws.cell(r, 4).value
        kl_da_xep = ws.cell(r, 5).value
        kl_da_do_kho = ws.cell(r, 6).value
        con_lai = ws.cell(r, 7).value
        forecast_end = ws.cell(r, 8).value
        if kl_giao is not None:
            kl_da_xep_v = float(kl_da_xep) if kl_da_xep is not None else 0.0
            kl_da_do_kho_v = float(kl_da_do_kho) if kl_da_do_kho is not None else 0.0
            con_lai_v = float(con_lai) if con_lai is not None else max(float(kl_giao) - kl_da_do_kho_v, 0.0)
            rows.append({
                "name": str(name).strip(),
                "start": fmt_date(start),
                "end": fmt_date(forecast_end),
                "klGiao": round(float(kl_giao), 2),
                "klDaXep": round(kl_da_xep_v, 2),
                "klDaDoKho": round(kl_da_do_kho_v, 2),
                "conLai": round(con_lai_v, 2),
            })
        r += 1
    return rows


def extract_po_ships(ws):
    """Sheet 'Kế hoạch tàu PO trong tháng': ke hoach cac chuyen tau OG du kien
    dat PO trong thang. Cot: B=ten tau, C=khoi luong PO, D=thoi gian dat PO,
    E=hop dong, F=OG so, G=thoi gian du kien giao hang, H=thoi gian du kien
    hoan thanh."""
    rows = []
    r = 2
    empty_streak = 0
    while empty_streak < 15 and r < 300:
        name = ws.cell(r, 2).value
        if name is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        kl_po = ws.cell(r, 3).value
        dat_po = ws.cell(r, 4).value
        contract = ws.cell(r, 5).value
        og_so = ws.cell(r, 6).value
        eta_giao = ws.cell(r, 7).value
        eta_hoan_thanh = ws.cell(r, 8).value
        if kl_po is not None:
            rows.append({
                "name": str(name).strip(),
                "klPO": round(float(kl_po), 2),
                "datPO": fmt_date(dat_po),
                "contract": str(contract).strip() if contract else "",
                "ogCode": str(og_so).strip() if og_so else "",
                "etaGiao": fmt_date(eta_giao),
                "etaHoanThanh": fmt_date(eta_hoan_thanh),
            })
        r += 1
    return rows


_LO_RE = re.compile(r"^\s*L[oôo]\s*(\d+)\s*:\s*(.+?)\s*$", re.IGNORECASE)


def extract_hd_detail_sheet(ws):
    """Sheet CHI TIET 1 hop dong (HĐ 17 / HĐ 24 / HĐ 25 / HĐ 26...): moi sheet co the co NHIEU
    khoi "Lô" (dot boc do) noi tiep nhau, moi khoi gom: 1 dong tieu de "Lô 0X : <ten tau me OG>"
    (cot B), 1 dong "Khối lượng hàng" ghi tong KL cua lo (cot B nhan, so luong o cot D), 1 dong
    tieu de cot bat dau bang "STT" (cot B), roi cac dong du lieu tung tau TS/SB (STT so o cot B)
    cho toi khi gap dong trong hoac khoi "Lô" ke tiep.

    QUAN TRONG: parser nay QUET THEO NOI DUNG O COT B (khong dua vao so dong co dinh), vi cac
    sheet lech nhau 1 dong (co sheet co dong trong dem giua tieu de va "Khối lượng hàng", co sheet
    khong) - da kiem chung thuc te tren du lieu that co it nhat 2 kieu bo tri khac nhau.

    Bo qua CAC DONG BI AN (row_dimensions[r].hidden) - theo dung quy uoc ghi trong dong 1 cua cac
    sheet nay ("LÔ NÀO ĐÃ BỐC DỠ XONG HOÀN TOÀN, ĐỀ NGHỊ ANH EM DDC CHO ẨN CÁC HÀNG CỦA LÔ ĐÓ"):
    cac lo/dong da an la lo cu da xong hoan toan, khong can hien thi lai tren dashboard."""
    lots = []
    cur = None
    in_data = False
    r = 1
    empty_streak = 0
    max_r = ws.max_row or 500
    while r <= max_r and empty_streak < 40:
        try:
            hidden = bool(ws.row_dimensions[r].hidden)
        except Exception:
            hidden = False
        if hidden:
            r += 1
            continue
        b = ws.cell(r, 2).value
        if b is None:
            empty_streak += 1
            # LUU Y: KHONG tat in_data o day. Trong 1 "Lô" co the co nhieu nhom tau (vd nhom
            # tau TS lon roi den nhom sa lan/tug nho) duoc ngan cach boi 1 dong trong va STT
            # danh so lai tu 1, nhung KHONG co dong tieu de "Lô .../STT" moi o giua - neu tat
            # in_data tai day se lam mat toan bo nhom tau thu 2 tro di cua lo. in_data chi tat
            # khi gap dong tieu de "Lô X:" moi (xem nhanh _LO_RE ben duoi).
            r += 1
            continue
        empty_streak = 0
        b_str = str(b).strip()
        m = _LO_RE.match(b_str)
        if m:
            cur = {"lot": m.group(1), "vessel": m.group(2), "totalQty": None, "rows": []}
            lots.append(cur)
            in_data = False
            r += 1
            continue
        if b_str.startswith("Khối lượng hàng") and cur is not None:
            d = ws.cell(r, 4).value
            try:
                cur["totalQty"] = round(float(d), 2) if d is not None else None
            except Exception:
                cur["totalQty"] = None
            in_data = False
            r += 1
            continue
        if b_str == "STT":
            in_data = True
            r += 1
            continue
        if in_data and cur is not None:
            name = ws.cell(r, 3).value
            if name:
                kl_nor = ws.cell(r, 6).value
                kl_boc = ws.cell(r, 7).value
                cur["rows"].append({
                    "name": str(name).strip(),
                    "nor": str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else "",
                    "port": str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else "",
                    "klNor": round(float(kl_nor), 2) if kl_nor is not None else None,
                    "klUnloaded": round(float(kl_boc), 2) if kl_boc is not None else None,
                    "start": str(ws.cell(r, 8).value).strip() if ws.cell(r, 8).value else "",
                    "end": str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else "",
                    "status": str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else "",
                })
        r += 1
    # Bo cac lo khong co dong du lieu nao (vd lo bi an toan bo, hoac khoi tieu de mo coi khong co
    # bang di kem) - tranh hien the rong tren dashboard.
    return [lot for lot in lots if lot["rows"]]


def main():
    try:
        import openpyxl
    except ImportError:
        die("chua cai thu vien openpyxl. Chay lenh sau roi thu lai:\n       pip install openpyxl")

    if not os.path.exists(EXCEL_PATH):
        die("khong tim thay file Excel:\n       " + EXCEL_PATH)
    if not os.path.exists(HTML_PATH):
        die("khong tim thay file:\n       " + HTML_PATH)

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        die("khong mo duoc file Excel (co the dang mo bang Excel tren may, hay dong file lai): " + str(e))

    for sheet_name in (SHEET_STOCK, SHEET_PRODUCTION):
        if sheet_name not in wb.sheetnames:
            die("khong tim thay sheet '%s' trong file Excel. Cac sheet hien co: %s" % (sheet_name, ", ".join(wb.sheetnames)))

    stock_rows = extract_stock_rows(wb[SHEET_STOCK])
    if not stock_rows:
        die("khong trich xuat duoc lo than nao dang ton tai bai tu sheet '%s'." % SHEET_STOCK)

    pile_stock_rows = extract_pile_stock_rows(wb[SHEET_STOCK])
    if not pile_stock_rows:
        die("khong trich xuat duoc lo than nao (pileStock) tu sheet '%s'." % SHEET_STOCK)

    report_date, month, year, month_days = extract_month_days(wb[SHEET_PRODUCTION])
    if not month_days:
        die("khong tim thay du lieu ton kho theo ngay trong sheet '%s'." % SHEET_PRODUCTION)

    def optional_sheet(sheet_key, extract_fn, label):
        if sheet_key in wb.sheetnames:
            return extract_fn(wb[sheet_key])
        log("CANH BAO: khong tim thay sheet '%s' - bo qua phan %s (cac slide/bang lien quan se hien trong)." % (sheet_key, label))
        return []

    og_indo_vtau = optional_sheet(SHEET_OG_INDO, extract_og_indo_vtau, "tau me OG (vi tri + tien do)")
    sb_vessels = optional_sheet(SHEET_SB, extract_sb_vessels, "tau con/sa lan (SB)")
    hd_tracking = optional_sheet(SHEET_HD_TRACKING, extract_hd_tracking, "theo doi hop dong giao nhan")
    po_ships = optional_sheet(SHEET_PO, extract_po_ships, "ke hoach tau PO trong thang")

    # 4 sheet chi tiet tung hop dong (HĐ 17/24/25/26) - dung rieng try/except QUANH TUNG SHEET (khong
    # chi optional_sheet ben tren) vi day la parser QUET NOI DUNG phuc tap hon han cac ham khac; neu
    # 1 sheet bi doi cau truc bat ngo trong tuong lai gay loi, CHI sheet do bi bo qua (rong []), 3
    # sheet con lai + toan bo dashboard van chay binh thuong, khong lam dung ca script.
    hd_contracts = {}
    for label, sheet_key in SHEET_HD_DETAIL.items():
        if sheet_key not in wb.sheetnames:
            log("CANH BAO: khong tim thay sheet '%s' - bo qua bang chi tiet %s tren Slide 05." % (sheet_key, label))
            hd_contracts[label] = []
            continue
        try:
            hd_contracts[label] = extract_hd_detail_sheet(wb[sheet_key])
        except Exception as e:
            log("CANH BAO: loi doc sheet '%s' (%s) - bo qua bang chi tiet %s tren Slide 05." % (sheet_key, str(e), label))
            hd_contracts[label] = []

    raw_data = {
        "reportDate": report_date,
        "month": month,
        "year": year,
        "monthDays": month_days,
        "sbVessels": sb_vessels,
        "stockRows": stock_rows,
        "pileStock": pile_stock_rows,
        "ogIndonesiaVtau": og_indo_vtau,
        "hdTracking": hd_tracking,
        "poShips": po_ships,
        "hdContracts": hd_contracts,
    }

    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # QUAN TRONG: khong gia dinh so khoang trang truoc "};" (ban goc viet tay thut 2 dau
    # cach, nhung json.dumps() lai xuat dau "}" dong ngoai cung KHONG thut dau) — neu quy
    # dinh cung so khoang trang, lan chay thu 2 tro di se khop nham qua ca cac khoi
    # "var yardOrder / segMeters / yardSegments" phia sau va xoa mat chung.
    # Noi dung JSON ben trong khong bao gio chua chuoi "};" (JSON chi dung dau phay/dong
    # ngoac don thuan, khong co dau cham phay) nen "};" dau tien gap duoc chinh la diem
    # ket thuc that su cua "var RAW_DATA = {...};", bat ke thut dau the nao.
    pattern = re.compile(r"var RAW_DATA = \{.*?\};", re.DOTALL)
    if not pattern.search(html):
        die("khong tim thay khoi 'var RAW_DATA' trong index.html — cau truc file mau co the da bi thay doi. "
            "Khong ghi de de tranh lam hong dashboard, vui long kiem tra lai thu cong.")

    new_block = "var RAW_DATA = " + json.dumps(raw_data, ensure_ascii=False, indent=2) + ";"
    # dung ham (khong phai chuoi) lam repl de tranh re.sub dien giai \1, \g<...> trong noi dung JSON
    new_html = pattern.sub(lambda m: new_block, html, count=1)

    # kiem tra nhanh: so luong the <script> khong doi (khong lam hong HTML)
    if html.count("<script>") != new_html.count("<script>"):
        die("phat hien bat thuong sau khi ghi du lieu moi — dung lai, khong luu file.")

    # kiem tra an toan bo sung: cac khai bao quan trong ngay sau RAW_DATA phai con nguyen
    # (phong truong hop pattern lo khop qua tay trong tuong lai neu cau truc file doi khac)
    for must_have in ("var yardOrder", "var segMeters", "var yardSegments", "function computeReport"):
        if must_have in html and must_have not in new_html:
            die("phat hien mat noi dung '%s' sau khi thay RAW_DATA — dung lai, khong luu file "
                "de tranh lam hong dashboard. Vui long bao lai de kiem tra generate_dashboard.py." % must_have)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)

    hd_contracts_lot_count = sum(len(v) for v in hd_contracts.values())
    log("OK: da cap nhat index.html tu du lieu ngay %s — %d lo than dang ton tai bai, "
        "%d lo than (pileStock) cho slide 3D moi, %d ngay du lieu trong thang %d/%d, "
        "%d tau OG Indonesia-VTau, %d tau con/sa lan (SB), %d hop dong theo doi giao nhan, "
        "%d tau PO ke hoach, %d lo dang hien (chua an) trong 4 hop dong HĐ 17/24/25/26." % (
            report_date, len(stock_rows), len(pile_stock_rows), len(month_days), month, year,
            len(og_indo_vtau), len(sb_vessels), len(hd_tracking), len(po_ships),
            hd_contracts_lot_count))


if __name__ == "__main__":
    main()
